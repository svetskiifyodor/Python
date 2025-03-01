import openpyxl
from openpyxl.styles import Alignment, Font
import pandas as pd

# Данные сотрудников
data = [
    {"Таб.номер": "0002", "Фамилия": "Петров П.П.", "Отдел": "Бухгалтерия", "Оклад, руб.": 3913.04,
     "Надбавки, руб.": 2608.7},
    {"Таб.номер": "0005", "Фамилия": "Васин В.В.", "Отдел": "Бухгалтерия", "Оклад, руб.": 5934.78,
     "Надбавки, руб.": 913.04},
    {"Таб.номер": "0001", "Фамилия": "Иванов И.И.", "Отдел": "Отдел кадров", "Оклад, руб.": 6000.00,
     "Надбавки, руб.": 4000.00},
    {"Таб.номер": "0003", "Фамилия": "Сидоров С.С.", "Отдел": "Отдел кадров", "Оклад, руб.": 5000.00,
     "Надбавки, руб.": 4500.00},
    {"Таб.номер": "0006", "Фамилия": "Львов Л.Л.", "Отдел": "Отдел кадров", "Оклад, руб.": 4074.07,
     "Надбавки, руб.": 2444.44},
    {"Таб.номер": "0007", "Фамилия": "Волков В.В.", "Отдел": "Отдел кадров", "Оклад, руб.": 1434.78,
     "Надбавки, руб.": 1434.78},
    {"Таб.номер": "0004", "Фамилия": "Мишин М.М.", "Отдел": "Столовая", "Оклад, руб.": 5500.00,
     "Надбавки, руб.": 3500.00}
]


# Функция для вычисления НДФЛ и суммы к выдаче
def calculate_salary(row):
    salary = row["Оклад, руб."] + row["Надбавки, руб."]
    ndfl = salary * 0.13
    net_salary = salary - ndfl
    return salary, ndfl, net_salary


# Функция для форматирования числа
def format_currency(value):
    return f"{value:,.2f}".replace(",", " ").replace(".", ",") + "р."


# Создаем книгу Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Зарплаты"

# Добавляем заголовки
headers = ["Таб.номер", "Фамилия", "Отдел", "Сумма по окладу, руб.", "Сумма по надбавками, руб.",
           "Сумма зарплаты, руб.", "НДФЛ, %", "Сумма НДФЛ, %", "Сумма к выдаче, руб."]
ws.append(headers)


# Функция для вычисления итоговых данных
def calculate_totals(department_data):
    total_salary = sum(row["Оклад, руб."] + row["Надбавки, руб."] for row in department_data)
    total_ndfl = sum((row["Оклад, руб."] + row["Надбавки, руб."]) * 0.13 for row in department_data)
    total_net_salary = total_salary - total_ndfl
    return total_salary, total_ndfl, total_net_salary


# Добавляем данные сотрудников и итоговые строки для каждого отдела
departments = ["Бухгалтерия", "Отдел кадров", "Столовая"]
for department in departments:
    department_data = [entry for entry in data if entry["Отдел"] == department]

    # Заполняем данные сотрудников отдела
    for entry in department_data:
        salary, ndfl, net_salary = calculate_salary(entry)
        row = [
            entry["Таб.номер"],
            entry["Фамилия"],
            entry["Отдел"],
            format_currency(entry["Оклад, руб."]),
            format_currency(entry["Надбавки, руб."]),
            format_currency(salary),
            "13%",
            format_currency(ndfl),
            format_currency(net_salary)
        ]
        ws.append(row)

    # Итоги по отделу
    total_salary, total_ndfl, total_net_salary = calculate_totals(department_data)
    row = [None, None, f"{department} Итог", format_currency(total_salary), format_currency(total_salary - total_ndfl),
           format_currency(total_salary), None, format_currency(total_ndfl), format_currency(total_net_salary)]
    ws.append(row)

    # Проверяем, если в строке содержится "Итог", то выделяем жирным
    for cell in ws[ws.max_row]:
        if "Итог" in str(cell.value):
            cell.font = Font(bold=True)

# Добавляем общий итог
total_salary, total_ndfl, total_net_salary = calculate_totals(data)
row = [None, None, "Общий Итог", format_currency(total_salary), format_currency(total_salary - total_ndfl),
       format_currency(total_salary), None, format_currency(total_ndfl), format_currency(total_net_salary)]
ws.append(row)

# Проверяем, если в строке содержится "Итог", то выделяем жирным
for cell in ws[ws.max_row]:
    if "Итог" in str(cell.value):
        cell.font = Font(bold=True)

# Настройка форматирования: повернуть текст в заголовках (без жирного шрифта)
for cell in ws[1]:
    cell.alignment = Alignment(textRotation=90, horizontal='center', vertical='center')

# Сохраняем файл
wb.save("salary_report.xlsx")
