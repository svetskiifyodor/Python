import openpyxl
import pandas as pd

# Функция для очистки строки от символов и преобразования в число
def parse_salary(salary_str):
    if isinstance(salary_str, str):
        # Убираем пробелы и символы "р." и заменяем запятую на точку
        salary_str = salary_str.replace("р.", "").replace(" ", "").replace(",", ".")
        try:
            return float(salary_str)
        except ValueError:
            return None
    return salary_str

# Загружаем существующий файл Excel, созданный в задании 1
wb = openpyxl.load_workbook("salary_report.xlsx")
ws = wb.active

# Извлекаем данные из таблицы в pandas DataFrame для удобства обработки
data = []
for row in ws.iter_rows(min_row=2, min_col=1, max_col=9, values_only=True):
    # Пропускаем строки с итогами (их нужно вычислять отдельно)
    if row[0] is not None and "Итог" not in str(row[2]):
        data.append({
            "Таб.номер": row[0],
            "Фамилия": row[1],
            "Отдел": row[2],
            "Сумма по окладу, руб.": row[3],
            "Сумма по надбавками, руб.": row[4],
            "Сумма зарплаты, руб.": parse_salary(row[5]),
            "НДФЛ, %": row[6],
            "Сумма НДФЛ, %": parse_salary(row[7]),
            "Сумма к выдаче, руб.": parse_salary(row[8])
        })

df = pd.DataFrame(data)

# 1. Человек с максимальной и минимальной зарплатой
max_salary_person = df.loc[df["Сумма зарплаты, руб."].idxmax()]
min_salary_person = df.loc[df["Сумма зарплаты, руб."].idxmin()]

# 2. Средняя зарплата по отделам
avg_salary_by_department = df.groupby("Отдел")["Сумма зарплаты, руб."].mean().reset_index()

# Выводим данные начиная с 14 строки и второго столбца
start_row = 14
start_col = 2

# Выводим информацию о максимальной и минимальной зарплате
ws.cell(row=start_row, column=start_col, value="Человек с максимальной зарплатой:")
ws.cell(row=start_row + 1, column=start_col, value=f"Фамилия: {max_salary_person['Фамилия']}")
ws.cell(row=start_row + 2, column=start_col, value=f"Зарплата: {max_salary_person['Сумма зарплаты, руб.']:,.2f}".replace(",", " ").replace(".", ","))

ws.cell(row=start_row + 3, column=start_col, value="Человек с минимальной зарплатой:")
ws.cell(row=start_row + 4, column=start_col, value=f"Фамилия: {min_salary_person['Фамилия']}")
ws.cell(row=start_row + 5, column=start_col, value=f"Зарплата: {min_salary_person['Сумма зарплаты, руб.']:,.2f}".replace(",", " ").replace(".", ","))

# Выводим среднюю зарплату по отделам
ws.cell(row=start_row + 7, column=start_col, value="Средняя зарплата по отделам:")
row_offset = start_row + 8
for idx, row in avg_salary_by_department.iterrows():
    ws.cell(row=row_offset, column=start_col, value=row["Отдел"])
    ws.cell(row=row_offset, column=start_col + 1, value=f"{row['Сумма зарплаты, руб.']:,.2f}".replace(",", " ").replace(".", ",") + "р.")
    row_offset += 1

# Сохраняем изменения в файл
wb.save("salary_report_updated.xlsx")
